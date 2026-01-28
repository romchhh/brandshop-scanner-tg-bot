import tempfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def calculate_statistics(results, inventory_data, art_map):
    """
    Підраховує статистику для повідомлення та Excel файлу
    Повертає словник з статистикою: total_sizes, matched_sizes, missing_sizes, not_scanned_sizes
    """
    # Загальна кількість розмірів (не артикулів, а саме розмірів)
    # Для товарів з розмірами - рахуємо розміри, для товарів без розмірів - рахуємо amount
    total_sizes = 0
    for norm_art, data_info in inventory_data.items():
        if data_info.get('sizes'):
            # Товар з розмірами
            for size, qty in data_info['sizes'].items():
                if size:  # Пропускаємо порожній розмір
                    total_sizes += qty
        else:
            # Товар без розмірів - рахуємо amount
            total_sizes += data_info.get('amount', 0)
    
    # Кількість співпадінь (сума кількостей розмірів для артикулів, які співпали)
    matched_sizes = 0
    for original_art in results.get('matched', []):
        if original_art in art_map:
            data = art_map[original_art]
            if data.get('sizes'):
                # Товар з розмірами
                for size, qty in data['sizes'].items():
                    if size:  # Пропускаємо порожній розмір
                        matched_sizes += qty
            else:
                # Товар без розмірів - рахуємо amount
                matched_sizes += data.get('amount', 0)
    
    # Кількість недостач (підраховуємо на основі артикулів з недостачею)
    missing_sizes = 0
    for original_art, missing_list in results.get('missing_sizes', {}).items():
        if original_art in art_map:
            data = art_map[original_art]
            # Для товарів без розмірів недостача вказується як "Недостача N шт"
            if not data.get('sizes') and data.get('amount', 0) > 0:
                # Товар без розмірів - парсимо "Недостача N шт"
                for missing_item in missing_list:
                    missing_str = str(missing_item)
                    import re
                    match = re.search(r'Недостача (\d+)', missing_str)
                    if match:
                        missing_sizes += int(match.group(1))
                    else:
                        # Якщо не знайдено формат, додаємо 1
                        missing_sizes += 1
            else:
                # Товар з розмірами - парсимо список недостач
                for missing_item in missing_list:
                    missing_str = str(missing_item)
                    # Перевіряємо формат "розмір (потрібно ще N)"
                    import re
                    match = re.search(r'\(потрібно ще (\d+)\)', missing_str)
                    if match:
                        # Якщо є кількість в дужках, використовуємо її
                        missing_sizes += int(match.group(1))
                    else:
                        # Якщо просто розмір без кількості, додаємо 1
                        missing_sizes += 1
    
    # Кількість не відсканованих (сума кількостей розмірів для артикулів, які не були відскановані)
    not_scanned_sizes = 0
    for art, sizes in results.get('not_scanned', {}).items():
        # sizes - це dict {розмір: кількість} або для товарів без розмірів {'': кількість}
        if sizes:
            # Перевіряємо, чи це товар без розмірів (порожній ключ)
            if '' in sizes:
                not_scanned_sizes += sizes['']
            else:
                # Товар з розмірами
                for size, qty in sizes.items():
                    if size:  # Пропускаємо порожній розмір
                        not_scanned_sizes += qty
    
    return {
        'total_sizes': total_sizes,
        'matched_sizes': matched_sizes,
        'missing_sizes': missing_sizes,
        'not_scanned_sizes': not_scanned_sizes
    }


def generate_inventory_excel(results, inventory_data, category_ua):
    """
    Генерує Excel файл з результатами перевірки
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Переоблік"
    
    # Заголовки
    headers = ["Арт", "Розміри", "Недостача", "Товар +"]
    ws.append(headers)
    
    # Стилі для заголовків
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Збираємо дані для файлу
    file_rows = []
    
    # Створюємо мапу для швидкого пошуку по original_art (бо в results ключі - це original_art)
    art_map = {}
    for norm_art, data_info in inventory_data.items():
        original_art = data_info['original_art']
        # sizes тепер це dict {нормалізований_розмір: кількість}
        # original_sizes це dict {нормалізований_розмір: оригінальний_розмір}
        art_map[original_art] = {
            'original_art': original_art,
            'sizes': data_info['sizes'],  # {нормалізований_розмір: кількість}
            'original_sizes': data_info.get('original_sizes', {}),  # {нормалізований_розмір: оригінальний_розмір}
            'amount': data_info.get('amount', 0)  # Загальна кількість для товарів без розмірів
        }
    
    def get_original_size(art_data, normalized_size):
        """Повертає оригінальний розмір з CSV, якщо він є, інакше нормалізований"""
        if normalized_size and 'original_sizes' in art_data:
            return art_data['original_sizes'].get(normalized_size, normalized_size)
        return normalized_size
    
    def format_sizes_for_display(data):
        """Форматує розміри для відображення. Для товарів без розмірів показує кількість"""
        # Якщо немає розмірів, але є amount - це товар без розмірів
        if not data.get('sizes') and 'amount' in data and data['amount'] > 0:
            return f"{data['amount']} шт"
        
        # Якщо є розміри
        if data.get('sizes'):
            sizes_list = []
            for normalized_size, qty in sorted(data['sizes'].items()):
                if normalized_size:  # Пропускаємо порожній розмір
                    # Використовуємо оригінальний розмір з CSV замість нормалізованого
                    original_size = get_original_size(data, normalized_size)
                    sizes_list.append(f"{original_size} ({qty})" if qty > 1 else original_size)
            if sizes_list:
                return ', '.join(sizes_list)
        
        return 'Без розмірів'
    
    def replace_normalized_sizes_in_text(text, art_data):
        """Замінює нормалізовані розміри на оригінальні в тексті"""
        if not text or not art_data or 'original_sizes' not in art_data:
            return text
        
        result = text
        original_sizes = art_data['original_sizes']
        
        # Замінюємо кожен нормалізований розмір на оригінальний
        for normalized_size, original_size in original_sizes.items():
            if normalized_size != original_size:
                # Замінюємо тільки якщо це окремий розмір (не частина іншого слова)
                import re
                pattern = r'\b' + re.escape(normalized_size) + r'\b'
                result = re.sub(pattern, original_size, result)
        
        return result
    
    # 1. Рядки з надлишком (зелені) - на початку
    for original_art, sizes in results.get('extra_sizes', {}).items():
        if original_art in art_map:
            data = art_map[original_art]
            # sizes тепер це dict {нормалізований_розмір: кількість}
            if data['sizes']:
                sizes_list = []
                for normalized_size, qty in sorted(data['sizes'].items()):
                    if normalized_size:  # Пропускаємо порожній розмір
                        # Використовуємо оригінальний розмір з CSV замість нормалізованого
                        original_size = get_original_size(data, normalized_size)
                        sizes_list.append(f"{original_size} ({qty})" if qty > 1 else original_size)
                sizes_str = ', '.join(sizes_list) if sizes_list else 'Без розмірів'
            else:
                sizes_str = 'Без розмірів'
            # sizes - це список рядків з описом надлишку (містить нормалізовані розміри)
            if isinstance(sizes, list):
                extra_str = ', '.join(sizes) if sizes else ''
            else:
                extra_str = ', '.join(sorted(sizes)) if sizes else ''
            # Замінюємо нормалізовані розміри на оригінальні
            extra_str = replace_normalized_sizes_in_text(extra_str, data)
            file_rows.append({
                'art': data['original_art'],
                'sizes': sizes_str,
                'missing': '',
                'extra': extra_str,
                'type': 'extra'  # Зелений
            })
    
    # 2. Рядки зі співпадінням (білі) - посередині
    for original_art in results.get('matched', []):
        if original_art in art_map:
            data = art_map[original_art]
            sizes_str = format_sizes_for_display(data)
            file_rows.append({
                'art': data['original_art'],
                'sizes': sizes_str,
                'missing': '',
                'extra': '',
                'type': 'matched'  # Білий
            })
    
    # 3. Рядки не відскановано (жовті) - після співпадінь
    for art, sizes in results.get('not_scanned', {}).items():
        # sizes - це dict {розмір: кількість} або для товарів без розмірів {'': кількість}
        if sizes:
            # Перевіряємо, чи це товар без розмірів (порожній ключ)
            if '' in sizes:
                sizes_str = f"{sizes['']} шт"
            else:
                sizes_list = []
                for size, qty in sorted(sizes.items()):
                    if size:  # Пропускаємо порожній розмір
                        sizes_list.append(f"{size} ({qty})" if qty > 1 else size)
                sizes_str = ', '.join(sizes_list) if sizes_list else 'Без розмірів'
        else:
            sizes_str = 'Без розмірів'
        
        file_rows.append({
            'art': art,
            'sizes': sizes_str,
            'missing': 'Не відскановано',
            'extra': '',
            'type': 'not_scanned'  # Жовтий
        })
    
    # 4. Рядки з недостачею (червоні) - в кінці
    for original_art, sizes in results.get('missing_sizes', {}).items():
        if original_art in art_map:
            data = art_map[original_art]
            sizes_str = format_sizes_for_display(data)
            # sizes - це список рядків з описом недостачі (містить нормалізовані розміри)
            if isinstance(sizes, list):
                missing_str = ', '.join(sizes) if sizes else ''
            else:
                missing_str = ', '.join(sorted(sizes)) if sizes else ''
            # Замінюємо нормалізовані розміри на оригінальні
            missing_str = replace_normalized_sizes_in_text(missing_str, data)
            file_rows.append({
                'art': data['original_art'],
                'sizes': sizes_str,
                'missing': missing_str,
                'extra': '',
                'type': 'missing'  # Червоний
            })
    
    # 5. Рядки не знайдені (червоні) - в кінці
    for art in results.get('not_found', []):
        # Шукаємо в inventory_data за оригінальним артикулом
        sizes_str = ''
        for norm_art, data_info in inventory_data.items():
            if data_info['original_art'] == art:
                # Створюємо тимчасовий об'єкт для форматування
                temp_data = {
                    'sizes': data_info['sizes'],
                    'original_sizes': data_info.get('original_sizes', {}),
                    'amount': data_info.get('amount', 0)
                }
                sizes_str = format_sizes_for_display(temp_data)
                break
        
        file_rows.append({
            'art': art,
            'sizes': sizes_str,
            'missing': 'Не знайдено в таблицях',
            'extra': '',
            'type': 'not_found'  # Червоний
        })
    
    # Перевіряємо, чи всі артикули з файлу додані до Excel
    # Збираємо всі артикули, які вже додані
    added_arts = set()
    for row in file_rows:
        added_arts.add(row['art'])
    
    # Додаємо артикули, які не потрапили в жодну категорію (на всяк випадок)
    for norm_art, data_info in inventory_data.items():
        original_art = data_info['original_art']
        if original_art not in added_arts:
            # Якщо артикул не доданий, додаємо його як співпадіння (білий)
            # Створюємо тимчасовий об'єкт для форматування
            temp_data = {
                'sizes': data_info['sizes'],
                'original_sizes': data_info.get('original_sizes', {}),
                'amount': data_info.get('amount', 0)
            }
            sizes_str = format_sizes_for_display(temp_data)
            file_rows.append({
                'art': original_art,
                'sizes': sizes_str,
                'missing': '',
                'extra': '',
                'type': 'matched'  # Білий - все нормально
            })
    
    # Додаємо рядки до таблиці
    for row_data in file_rows:
        ws.append([
            row_data['art'],
            row_data['sizes'],
            row_data['missing'],
            row_data.get('extra', '')  # Товар + (надлишок)
        ])
    
    # Застосовуємо форматування
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Світло-зелений
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Світло-червоний
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Білий
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Жовтий
    
    # Застосовуємо кольори до рядків
    for idx, row_data in enumerate(file_rows):
        row_num = idx + 2  # Починаємо з 2, бо 1 - заголовок
        row_type = row_data['type']
        
        if row_type == 'extra':
            fill = green_fill
        elif row_type in ['missing', 'not_found']:
            fill = red_fill
        elif row_type == 'not_scanned':
            fill = yellow_fill  # Жовтий для не відсканованих
        elif row_type == 'matched':
            fill = white_fill  # Явно встановлюємо білий для співпадінь
        else:
            fill = white_fill  # За замовчуванням білий
        
        # Застосовуємо кольори до всіх комірок рядка
        for col_num in range(1, 5):  # 4 стовпці
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Налаштування ширини стовпців
    ws.column_dimensions['A'].width = 20  # Арт
    ws.column_dimensions['B'].width = 30  # Розміри
    ws.column_dimensions['C'].width = 30  # Недостача
    ws.column_dimensions['D'].width = 30  # Товар +
    
    # Підрахунок статистики
    stats = calculate_statistics(results, inventory_data, art_map)
    total_sizes = stats['total_sizes']
    matched_sizes = stats['matched_sizes']
    missing_sizes = stats['missing_sizes']
    not_scanned_sizes = stats['not_scanned_sizes']
    
    # Додаємо порожній рядок перед статистикою
    ws.append([])
    
    # Додаємо рядок зі статистикою
    stats_row = ws.max_row + 1
    ws.cell(row=stats_row, column=1, value=f"Розмірів {total_sizes}")
    ws.cell(row=stats_row, column=1).font = Font(bold=True)
    ws.cell(row=stats_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    ws.cell(row=stats_row, column=2, value=f"Сошлося {matched_sizes}")
    ws.cell(row=stats_row, column=2).font = Font(bold=True)
    ws.cell(row=stats_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
    
    ws.cell(row=stats_row, column=3, value=f"Недостача {missing_sizes}")
    ws.cell(row=stats_row, column=3).font = Font(bold=True)
    ws.cell(row=stats_row, column=3).alignment = Alignment(horizontal="left", vertical="center")
    
    ws.cell(row=stats_row, column=4, value=f"Не відскановано {not_scanned_sizes}")
    ws.cell(row=stats_row, column=4).font = Font(bold=True)
    ws.cell(row=stats_row, column=4).alignment = Alignment(horizontal="left", vertical="center")
    
    # Застосовуємо білий фон для рядка статистики
    for col_num in range(1, 5):
        cell = ws.cell(row=stats_row, column=col_num)
        cell.fill = white_fill
    
    # Зберігаємо файл
    file_path = tempfile.mktemp(suffix='.xlsx')
    wb.save(file_path)
    
    return file_path
